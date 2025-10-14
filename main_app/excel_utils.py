"""
Excel export utilities for Citizens Secondary School COSA
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.utils import timezone
from datetime import datetime


def create_excel_response(filename, workbook):
    """Create HTTP response for Excel file download"""
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response


def style_header_cell(cell):
    """Apply styling to header cells"""
    cell.font = Font(bold=True, color="FFFFFF", size=12)
    cell.fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    cell.border = thin_border


def style_data_cell(cell):
    """Apply styling to data cells"""
    cell.font = Font(size=11)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    cell.border = thin_border


def export_alumni_to_excel(alumni_queryset, export_type="all"):
    """
    Export alumni data to Excel format

    Args:
        alumni_queryset: QuerySet of Alumni objects
        export_type: Type of export ("all", "basic", "detailed")

    Returns:
        HttpResponse with Excel file
    """
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "COSA Alumni Members"

    # Define headers based on export type
    if export_type == "basic":
        headers = [
            "Student ID", "Full Name", "Email", "Phone",
            "Completion Year", "Level", "Registration Date"
        ]
    elif export_type == "detailed":
        headers = [
            "Student ID", "First Name", "Last Name", "Email", "Phone",
            "Completion Year", "Level", "Department", "Current Job",
            "Company", "Location", "LinkedIn Profile", "Registration Date", "Last Updated"
        ]
    else:  # all
        headers = [
            "Student ID", "First Name", "Last Name", "Email", "Phone",
            "Completion Year", "Level", "Department", "Current Job",
            "Company", "Location", "LinkedIn Profile", "Bio", "Skills",
            "Registration Date", "Last Updated", "Status"
        ]

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num, value=header)
        style_header_cell(cell)

    # Write data
    row_num = 2
    for alumni in alumni_queryset:
        user = alumni.admin  # assuming FK to auth user is 'admin'

        # Derive common fields safely
        grad_year_val = getattr(getattr(alumni, 'graduation_year', None), 'year', '') or ''
        level_name = getattr(getattr(alumni, 'level', None), 'name', '') or ''
        department_obj = getattr(alumni, 'department', None)
        department_name = getattr(department_obj, 'name', '') if department_obj else ''
        current_company_obj = getattr(alumni, 'current_company', None)
        current_company_name = getattr(current_company_obj, 'name', '') if current_company_obj else ''
        location = ''
        if getattr(alumni, 'current_city', None):
            country = getattr(alumni, 'current_country', '') or ''
            location = f"{alumni.current_city}, {country}".strip().strip(', ')
        phone_val = getattr(user, 'phone', '') or ''

        if export_type == "basic":
            data = [
                getattr(alumni, 'student_id', ''),
                user.get_full_name(),
                user.email,
                phone_val,
                grad_year_val,
                level_name,
                alumni.created_at.strftime('%Y-%m-%d') if getattr(alumni, 'created_at', None) else ''
            ]
        elif export_type == "detailed":
            data = [
                getattr(alumni, 'student_id', ''),
                user.first_name,
                user.last_name,
                user.email,
                phone_val,
                grad_year_val,
                level_name,
                department_name,
                getattr(alumni, 'job_title', '') or getattr(alumni, 'current_job', '') or '',
                current_company_name or getattr(alumni, 'company', '') or '',
                location or getattr(alumni, 'location', '') or '',
                getattr(alumni, 'linkedin_profile', '') or '',
                alumni.created_at.strftime('%Y-%m-%d %H:%M') if getattr(alumni, 'created_at', None) else '',
                alumni.updated_at.strftime('%Y-%m-%d %H:%M') if getattr(alumni, 'updated_at', None) else ''
            ]
        else:  # all
            data = [
                getattr(alumni, 'student_id', ''),
                user.first_name,
                user.last_name,
                user.email,
                phone_val,
                grad_year_val,
                level_name,
                department_name,
                getattr(alumni, 'job_title', '') or getattr(alumni, 'current_job', '') or '',
                current_company_name or getattr(alumni, 'company', '') or '',
                location or getattr(alumni, 'location', '') or '',
                getattr(alumni, 'linkedin_profile', '') or '',
                getattr(alumni, 'bio', '') or '',
                getattr(alumni, 'skills', '') or '',
                alumni.created_at.strftime('%Y-%m-%d %H:%M') if getattr(alumni, 'created_at', None) else '',
                alumni.updated_at.strftime('%Y-%m-%d %H:%M') if getattr(alumni, 'updated_at', None) else '',
                'Active' if getattr(user, 'is_active', False) else 'Inactive'
            ]

        # Write row data
        for col_num, value in enumerate(data, 1):
            cell = worksheet.cell(row=row_num, column=col_num, value=value)
            style_data_cell(cell)

        row_num += 1

    # Auto-adjust column widths
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for row in worksheet[column_letter]:
            try:
                if len(str(row.value)) > max_length:
                    max_length = len(str(row.value))
            except Exception:
                pass
        adjusted_width = min(max(max_length + 2, 10), 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width

    # Add summary information
    summary_row = row_num + 2
    worksheet.cell(row=summary_row, column=1, value="Export Summary:")
    worksheet.cell(row=summary_row, column=1).font = Font(bold=True)

    worksheet.cell(row=summary_row + 1, column=1,
                   value=f"Total Registered Members Exported: {alumni_queryset.count()}")
    worksheet.cell(row=summary_row + 2, column=1,
                   value=f"Export Date: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}")
    worksheet.cell(row=summary_row + 3, column=1, value=f"Export Type: {export_type.title()}")
    worksheet.cell(row=summary_row + 4, column=1, value="Generated by Citizens Secondary School COSA System")

    # Generate filename
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"COSA_Alumni_Export_{export_type}_{timestamp}.xlsx"

    return create_excel_response(filename, workbook)


def export_alumni_by_graduation_year(graduation_year):
    """Export alumni by specific graduation year"""
    from .models import Alumni

    alumni_queryset = Alumni.objects.filter(
        graduation_year=graduation_year
    ).select_related('admin', 'graduation_year', 'level')

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = f"Class of {getattr(graduation_year, 'year', graduation_year)}"

    # Headers
    headers = [
        "Student ID", "Full Name", "Email", "Phone", "Level",
        "Department", "Current Job", "Company", "Location", "Registration Date"
    ]

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num, value=header)
        style_header_cell(cell)

    # Write data
    row_num = 2
    for alumni in alumni_queryset:
        user = alumni.admin
        level_name = getattr(getattr(alumni, 'level', None), 'name', '') or ''
        department_obj = getattr(alumni, 'department', None)
        department_name = getattr(department_obj, 'name', '') if department_obj else ''
        company_name = getattr(getattr(alumni, 'current_company', None), 'name', '') or getattr(alumni, 'company', '') or ''
        current_job = getattr(alumni, 'job_title', '') or getattr(alumni, 'current_job', '') or ''
        location = getattr(alumni, 'location', '')
        if not location and getattr(alumni, 'current_city', None):
            country = getattr(alumni, 'current_country', '') or ''
            location = f"{alumni.current_city}, {country}".strip().strip(', ')

        data = [
            getattr(alumni, 'student_id', ''),
            user.get_full_name(),
            user.email,
            getattr(user, 'phone', '') or '',
            level_name,
            department_name,
            current_job,
            company_name,
            location,
            alumni.created_at.strftime('%Y-%m-%d') if getattr(alumni, 'created_at', None) else ''
        ]

        for col_num, value in enumerate(data, 1):
            cell = worksheet.cell(row=row_num, column=col_num, value=value)
            style_data_cell(cell)

        row_num += 1

    # Auto-adjust column widths
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for row in worksheet[column_letter]:
            try:
                if len(str(row.value)) > max_length:
                    max_length = len(str(row.value))
            except Exception:
                pass
        adjusted_width = min(max(max_length + 2, 10), 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width

    # Add summary
    summary_row = row_num + 2
    worksheet.cell(row=summary_row, column=1,
                   value=f"Class of {getattr(graduation_year, 'year', graduation_year)} Alumni Export")
    worksheet.cell(row=summary_row, column=1).font = Font(bold=True)
    worksheet.cell(row=summary_row + 1, column=1,
                   value=f"Total Registered Members: {alumni_queryset.count()}")
    worksheet.cell(row=summary_row + 2, column=1,
                   value=f"Export Date: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}")

    filename = f"COSA_Class_of_{getattr(graduation_year, 'year', graduation_year)}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return create_excel_response(filename, workbook)


def export_alumni_statistics():
    """Export alumni statistics and analytics"""
    from .models import Alumni, GraduationYear, Level

    workbook = Workbook()

    # Statistics Sheet
    stats_sheet = workbook.active
    stats_sheet.title = "Alumni Statistics"

    # Headers
    headers = ["Metric", "Count", "Percentage"]
    for col_num, header in enumerate(headers, 1):
        cell = stats_sheet.cell(row=1, column=col_num, value=header)
        style_header_cell(cell)

    # Calculate statistics
    total_alumni = Alumni.objects.count()
    active_alumni = Alumni.objects.filter(admin__is_active=True).count()

    def pct(n, d):
        return f"{(n / d * 100):.1f}%" if d > 0 else "0%"

    stats_data = [
        ("Total Registered Members", total_alumni, "100%"),
        ("Active Alumni", active_alumni, pct(active_alumni, total_alumni)),
        ("Inactive Alumni", total_alumni - active_alumni, pct(total_alumni - active_alumni, total_alumni)),
    ]

    # Graduation year stats
    for grad_year in GraduationYear.objects.all().order_by('-year'):
        count = Alumni.objects.filter(graduation_year=grad_year).count()
        stats_data.append((f"Class of {grad_year.year}", count, pct(count, total_alumni)))

    # Level stats
    for level in Level.objects.all():
        count = Alumni.objects.filter(level=level).count()
        stats_data.append((f"{level.name} Level", count, pct(count, total_alumni)))

    # Write statistics
    row_num = 2
    for metric, count, percentage in stats_data:
        stats_sheet.cell(row=row_num, column=1, value=metric)
        stats_sheet.cell(row=row_num, column=2, value=count)
        stats_sheet.cell(row=row_num, column=3, value=percentage)

        for col_num in range(1, 4):
            style_data_cell(stats_sheet.cell(row=row_num, column=col_num))

        row_num += 1

    # Auto-adjust column widths
    for col_num in range(1, 4):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for row in stats_sheet[column_letter]:
            try:
                if len(str(row.value)) > max_length:
                    max_length = len(str(row.value))
            except Exception:
                pass
        adjusted_width = min(max(max_length + 2, 10), 50)
        stats_sheet.column_dimensions[column_letter].width = adjusted_width

    filename = f"COSA_Alumni_Statistics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return create_excel_response(filename, workbook)
